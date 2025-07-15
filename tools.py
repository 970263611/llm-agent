import os
import json
import subprocess
import logging
from datetime import datetime
from typing import Dict, Any
from langchain.tools import tool
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import docx

logger = logging.getLogger(__name__)

# =================工具函数 =================
class ProjectTools:
    # 类变量缓存工具实例
    tools_cache = None

    @classmethod
    def get_all_tools(cls) -> list:
        """获取所有工具实例"""
        if cls.tools_cache is None:
            cls.tools_cache = [
                cls.create_file,
                cls.run_command,
                # cls.get_directory_structure
            ]
            logger.debug("工具列表初始化完成")
        return cls.tools_cache
    
    @tool
    def create_file(file_path: str, content: str) -> Dict[str, Any]:
        """创建文件并写入内容"""
        try:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            result = {
                "status": "success",
                "message": f"文件创建成功: {file_path}",
                "file_path": os.path.abspath(file_path)
            }
            logger.info(f"工具调用成功: create_file {json.dumps(result, ensure_ascii=False)}")
            return result
        except Exception as e:
            error_msg = f"文件创建失败: {str(e)}"
            logger.error(f"工具调用失败: create_file {error_msg}")
            return {"status": "error", "message": error_msg}  
    @tool
    def run_command(command: str, path: str = ".") -> Dict[str, Any]:
        """在指定目录执行终端命令"""
        try:
            abs_path = os.path.abspath(path)
            if not os.path.exists(abs_path):
                os.makedirs(abs_path, exist_ok=True)

            # 在终端执行命令返回结果
            result = subprocess.run(
                command,
                cwd=abs_path,
                shell=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=15
            )

            def safe_decode(byte_data):
                if not byte_data: 
                    return None
                try:
                    return byte_data.decode('utf-8')
                except UnicodeDecodeError:
                     # 中文Windows默认编码
                    return byte_data.decode('gbk', errors='replace') 
            # 得到结果输出,适配不同的操作编码环境
            stdout = safe_decode(result.stdout)
            stderr = safe_decode(result.stderr)

            # 返回给tool_node节点
            output = {
                "status": "success" if result.returncode == 0 else "error",
                "command": command,
                "output": (stdout or "")[:500], 
                "error": (stderr or "")[:500] if result.returncode != 0 else None
            }
            
            logger.info(f"命令执行结果: {json.dumps(output, ensure_ascii=False)}")
            return output
        except Exception as e:
            error_msg = f"命令执行异常: {str(e)}"
            logger.error(f"命令执行失败: {error_msg}", exc_info=True)
            return {"status": "error", "message": error_msg}
        
    @tool
    def get_directory_structure(path: str = ".") -> Dict[str, Any]:
        """
        获取原始目录结构数据（不包含任何分析逻辑）
        返回包含完整目录树和文件信息的JSON结构
        """
        def scan_directory(directory: str, max_depth: int = 3, current_depth: int = 0) -> Dict:
            if current_depth >= max_depth:
                return {
                    "name": os.path.basename(directory),
                    "type": "directory",
                    "path": directory,
                    "note": f"目录深度超过{max_depth}层已折叠"
                }
            
            structure = {
                "name": os.path.basename(directory),
                "type": "directory",
                "path": directory,
                "items": []
            }
            
            try:
                for item in sorted(os.listdir(directory)):
                    item_path = os.path.join(directory, item)
                    if os.path.isdir(item_path):
                        structure["items"].append(
                            scan_directory(item_path, max_depth, current_depth + 1)
                        )
                    else:
                        structure["items"].append({
                            "name": item,
                            "type": "file",
                            "path": item_path,
                            "size": os.path.getsize(item_path),
                            "extension": os.path.splitext(item)[1].lower()
                        })
            except PermissionError:
                structure["error"] = "权限不足"
            
            return structure

        try:
            abs_path = os.path.abspath(path)
            return {
                "status": "success",
                "path": abs_path,
                "structure": scan_directory(abs_path),
                "timestamp": datetime.now().isoformat()
            }
        except Exception as e:
            return {
                "status": "error",
                "message": str(e)
            }  


    @classmethod
    def upload_and_analyze_file(cls, file_path: str, analysis_instructions: str = "请分析此文件内容") -> Dict[str, Any]:
        """
        上传并分析文件内容(PDF, Excel, Word, CSV等)，返回结构化数据供LLM分析。
        
        参数:
            file_path: 要分析的文件路径
            analysis_instructions: 对LLM的具体分析指令
            
        返回:
            包含文件内容摘要和元数据的字典，适合LLM进一步处理
        """
        try:
            # 验证文件存在
            if not os.path.exists(file_path):
                return {"status": "error", "message": "文件不存在"}
            
            # 获取文件信息
            file_name = os.path.basename(file_path)
            file_size = os.path.getsize(file_path)
            file_ext = os.path.splitext(file_name)[1].lower()
            
            # 根据文件类型解析内容
           
            content = ""
            if file_ext == '.pdf':
                content = cls.extract_pdf_content(file_path)
            elif file_ext in ('.xlsx', '.xls'):
                content = cls.extract_excel_content(file_path)
            elif file_ext == '.docx':
                content = cls.extract_word_content(file_path)
            elif file_ext == '.txt':
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
            else:
                return {"status": "error", "message": f"不支持的文件类型: {file_ext}"}
            
            return {
                "status": "success",
                "file_name": file_name,
                "file_size": file_size,
                "content_count": len(content),
                "content": content,  
                "instructions": analysis_instructions
            }
            
        except Exception as e:
            error_msg = f"文件分析失败: {str(e)}"
            logger.error(f"文件分析错误: {error_msg}", exc_info=True)
            return {"status": "error", "message": error_msg}
        
    @classmethod
    def extract_pdf_content(self,file_path: str) -> str:
        """提取PDF文件文本内容"""
        content = []
        with open(file_path, 'rb') as f:
            reader = PdfReader(f)
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    content.append(text)
        return "\n".join(content)
    
    @classmethod
    def extract_excel_content(self,file_path: str) -> str:
        """提取Excel文件内容"""
        content = []
        wb = load_workbook(filename=file_path, read_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            content.append(f"\n=== 工作表: {sheet_name} ===")
            for row in sheet.iter_rows(values_only=True):
                row_content = " | ".join(str(cell) if cell is not None else "" for cell in row)
                content.append(row_content)
        return "\n".join(content)
    
    @classmethod
    def extract_csv_content(self,file_path: str) -> str:
        """提取CSV文件内容"""
        try:
            # 尝试UTF-8编码
            df = pd.read_csv(file_path, encoding='utf-8')
        except UnicodeDecodeError:
            try:
                # 尝试GBK编码(中文常见)
                df = pd.read_csv(file_path, encoding='gbk')
            except Exception as e:
                raise ValueError(f"无法解析CSV文件编码: {str(e)}")
        
        return df.to_string(index=False)
    
    @classmethod
    def extract_word_content(self, file_path: str) -> str:
        """提取Word文档内容"""
        doc = docx.Document(file_path)
        return "\n".join(paragraph.text for paragraph in doc.paragraphs if paragraph.text) 

